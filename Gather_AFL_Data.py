# To do: Add Whether they won and margin (maybe negative numbers)
# add advanced stats & brownlow votes possibly?

#Author: Mitchell Gill
#Date: 03/06/2020 and continuing to update
#A range of functions to gather data from 'footywire'
#Including which teams are in which match ID's
#And a way of scraping for each teams stats for each game they play in after creating
#a txt of which teams play in which games

import sys
import requests
import pprint
import os.path
import xlsxwriter
import openpyxl
from openpyxl import Workbook, load_workbook
from os import path
from bs4 import BeautifulSoup

#gets key from dictionary
def get_key(val, my_dict):
    for key, value in my_dict.items():
         if val == value:
             return key

#Scrapes webpage for which teams played
#inputs are a team dictionary the team we are looking at and the match num
def scrape_match_teams_playing(teams, team_id, match_num):
    flag = 0
    team = teams.get(str(team_id))
    URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid=" + str(match_num)
    #print(URL)
    current_team = (teams[str(team_id)])
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
    #returns the teams playing for each match by getting the text element
    #of the HTML returned by the soup object
    data = [element.text for element in soup.find_all('td', class_='bnorm', width='190')]
    #print(data)
    #match_data = results.find_all('td', class_="statdata")
    if current_team in data:
        flag = 1
    return flag

#Scrapes webpage for the basic stats and returns an array of the data
#inputs the teams dict, current team
def scrape_match_basic_stats(teams, team_id):
    team = teams.get(str(team_id))
    f = open(team+"_data.txt", 'r')
    M_IDs = f.readlines()
    M_IDs = [x.rstrip() for x in M_IDs]
    count = 1
    for mn in M_IDs:
        #start the for loop here and keep track of line num
        URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid="+str(mn)
        print(URL)
        page = requests.get(URL)
        soup = BeautifulSoup(page.content, 'html.parser')
        stat_array = []
        data = [element.text for element in soup.find_all('td', class_='bnorm', width='190')]
        test_data = [element.text for element in soup.find_all('td', class_="lnorm", height='22')]
        #uses test_data to chop out which round the match is being played as
        round = determine_round(test_data[0])
        stat_array.append(round)
        #determines if the given team for the match is home or away
        home = 0
        #the home team appears first  in the data array
        if(team == data[0]):
            stat_array.append(0)
            oppo_ID = get_key(data[1], teams)
            stat_array.append(int(oppo_ID))
        #otherwise they must be the away team
        else:
            home = 1
            stat_array.append(1)
            oppo_ID = get_key(data[0], teams)
            stat_array.append(int(oppo_ID))
        stats_pulled = [element.text for element in soup.find_all('td', class_="statdata")]
        stat_array = wrangle_stats_basic(stats_pulled, stat_array, home)
        write_to_excel(team, stat_array, count)
        count = count + 1

#this is fucking disgusting
#takes the big soup of all HTML text code labelled as 'statdata'
#reverses the soup as the important stuff is at the back
#kicks is the 'last stat' so it counts to kicks + 1 for the final stat point
#splices all other data off and leaves us with the good stuff
#if the team we are getting stats for is the home team it starts from array = 0
#otherwise it starts from 2 and either way they do every third step
def wrangle_stats_basic(stats_pulled, stat_array, home):
    stats_pulled.reverse()
    i = 1
    for x in stats_pulled:
        if(x == "Kicks"):
            i = i + 1
            break
        i = i+1
    stats_pulled = stats_pulled[:i]
    stats_pulled.reverse()
    if(home == 0):
        for x in stats_pulled[::3]:
            #gets rid of % mark
            if("%" in x):
                x = x[:-1]
            #gets ride of kg or cm
            elif(("cm" in x) or ("kg" in x)):
                x = x[:-2]
            #gets rid of mth
            elif("mth" in x):
                x = x[:-3]
                #if mth is 10 or 11
                if(len(x) == 7):
                    y = x[-1]+x[-2]
                    #makes the mth = to decimal
                    z = float(y)*8.33
                    #don't think this would actually happen
                    if(z < 9):
                        z = str(z)
                        z = z[:1]
                        x = x[:2] + '.' + z
                    else:
                        z = str(z)
                        z = z[:2]
                        x = x[:2] + '.' + z
                #if mth is 0-9
                else:
                    y = x[-1]
                    z = float(y)*8.33
                    #will be less than 9 if mth is 0 or 1
                    #will result in turning 24yr 11mth to 24.91
                    if(z < 9):
                        z = str(z)
                        z = z[:1]
                        x = x[:2] + '.' + z
                    else:
                        z = str(z)
                        z = z[:2]
                        x = x[:2] + '.' + z
            stat_array.append(float(x))
    #does the same as above except the current team is the away teams
    #should really put this messy shit into a function but its too late
    else:
        for x in stats_pulled[2::3]:
            if("%" in x):
                x = x[:-1]
            elif(("cm" in x) or ("kg" in x)):
                x = x[:-2]
            elif("mth" in x):
                x = x[:-3]
                if(len(x) == 7):
                    y = x[-1]+x[-2]
                    z = float(y)*8.33
                    if(z < 9):
                        z = str(z)
                        z = z[:1]
                        x = x[:2] + '.' + z
                    else:
                        z = str(z)
                        z = z[:2]
                        x = x[:2] + '.' + z
                else:
                    y = x[-1]
                    z = float(y)*8.33
                    if(z < 9):
                        z = str(z)
                        z = z[:1]
                        x = x[:2] + '.' + z
                    else:
                        z = str(z)
                        z = z[:2]
                        x = x[:2] + '.' + z
            stat_array.append(float(x))
    return stat_array

#Example input 'Round 23, Marvel Stadium... etc' will look at either first character
# or comma position to determine what round it is
# finals rounds are given 25-28 values accordingly
def determine_round(round_string):
    round = 0
    if (round_string[0] == 'Q' or round_string[0] == 'E'):
        round = 25
    elif (round_string[0] == "S"):
        round = 26
    elif (round_string[0] == "P"):
        round = 27
    elif (round_string[0] == "G"):
        round = 28
    else:
        if(round_string[7] == ","):
            round = int(round_string[6])
        else:
            round = int(round_string[6] + round_string[7])
    return round

#opens an excel file based on the team name
#if the excel file doesn't exist it creates the excel file and adds the labeled column
#along with the first set of statistics
#otherwise it opens the existing file and adds the relevant stats into the next open column
def write_to_excel(team, stat_array, match_count):
    if(not(path.exists(team+'_stats.xlsx'))):
        wb = Workbook()
        ws = wb.active
        labels = ['Round', 'H/A?', 'Team_against_ID', 'Kicks', 'Handballs', 'Disposals',
        'Kick to HB Ratio', 'Marks', 'Tackles', 'Hitouts', 'Frees For', 'Frees Against',
        'Goals Kicked', 'Goal Assists', 'Behinds Kicked', 'Rushed Behinds', 'Scoring Shots',
        'Conversion %', 'Disposals Per Goal', 'Disposals Per Scoring Shot', 'Clearances',
        'Clangers', 'Rebound 50s', 'Inside 50s', 'In50s Per Scoring Shot', 'In50s Per Goal',
        '% In50s Score', '% In50s Goal', 'Height', 'Weight', 'Age', 'Av Games', '<50 Games',
        '50-99 Games', '100-149 Games', '>150 Games', 'G from Marks', 'G from FK', 'G from 50m Pen',
        'G from Play']
        i = 0
        j = 0
        #iterates through each column, in the given range, here it is the 1st column
        #goes to maximum rows of the length of labels for stats
        for col in ws.iter_cols(max_col=1, max_row=len(labels)):
            for cell in col:
                cell.value = labels[i]
                i = i+1
        #iterates to the column that should be free next
        for col in ws.iter_cols(min_col=match_count+1, max_col=match_count+1, max_row=len(stat_array)):
            for cell in col:
                cell.value = stat_array[j]
                j = j + 1
        wb.save(filename = team+'_stats.xlsx')
    else:
        wb = load_workbook(team+'_stats.xlsx')
        ws = wb.active
        j = 0
        for col in ws.iter_cols(min_col=match_count+1, max_col=match_count+1, max_row=len(stat_array)):
            for cell in col:
                cell.value = stat_array[j]
                j = j + 1
        wb.save(filename = team+'_stats.xlsx')


#Scrapes webpage for advanced stats match data for a given team and match
#def scrape_match_advanced_stats(teams, team_id, match_num):
#    team = teams.get(str(team_id))
#    URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid="+str(match_num)+"&advv=Y"
#    print(URL)
#    page = requests.get(URL)
    #print(page.content)
#    soup = BeautifulSoup(page.content, 'html.parser')

#creates a dictionary with each teams identifier on afl_tables
def createTeamDict():
    teamDict = {
    "1" : "Adelaide",
    "2" : "Brisbane",
    "3" : "Carlton",
    "4" : "Collingwood",
    "5" : "Essendon",
    "6" : "Fremantle",
    "7" : "Geelong",
    "8" : "Gold Coast",
    "9" : "GWS",
    "10": "Hawthorn",
    "11": "Melbourne",
    "12": "North Melbourne",
    "13": "Port Adelaide",
    "14": "Richmond",
    "15": "St Kilda",
    "16": "Sydney",
    "17": "West Coast",
    "18": "Western Bulldogs"
    }
    return teamDict

#creates a file for each team that has the ID's of each of their matches  since 2011
def createTeamMatchFile(team_int, team_dict):
    #gets current team we are looking at
    current_team = (team_dict[str(team_int)])
    print(current_team)
    textfile = open(current_team+"_data.txt", 'w+')
    match = 1
    #round 1 2011, start of the expansion teams
    p = 5147
    while(p<6362):
        #checks if current game has current team
        flag_num = scrape_match_teams_playing(team_dict, team_int, p)
        #if it does it records the match ID, makes it easier for the future
        if(flag_num == 1):
            #matches played because it doesn't take into account the bye round
            print("Match: "+str(match))
            match = match + 1
            textfile.write(str(p)+'\n')
        p = p+1
    #match starts at EF 2016 WC vs WB
    j = 9298
    #End of Round 1 2020
    while(j<9936):
        #checks if current game has current team
        flag_num = scrape_match_teams_playing(team_dict, team_int, j)
        #if it does it records the match ID, makes it easier for the future
        if(flag_num == 1):
            #matches played because it doesn't take into account the bye round
            print("match: "+str(match))
            match = match + 1
            textfile.write(str(j)+'\n')
        j = j+1
    textfile.close()

def main():
    teams = createTeamDict()
    #for each team do:
    scrape_match_basic_stats(teams,6)
    #i = 1
    #should go through each of the 18 teams and create a txt file of their match ID's
    #while(i<19):
    #    createTeamMatchFile(i,teams)
    #    i = i+1

if __name__ == '__main__':
    main()
