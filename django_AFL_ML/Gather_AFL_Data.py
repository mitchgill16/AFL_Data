# To do: add update functions for the txt and xlsx files
#you forgot game 9936, and R2 2020 onwards starts in like 10,000s

#Author: Mitchell Gill
#Date: 07/06/2020 and continuing to update
#A range of functions to gather data from 'footywire'
#Including which teams are in which match ID's
#And a way of scraping for each teams stats for each game they play in after creating
#a txt of which teams play in which games

class gatherer:

    import sys
    import requests
    import pprint
    import os.path
    import xlsxwriter
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from os import path
    from bs4 import BeautifulSoup

    #gets key from dictionary
    def get_key(self, val, my_dict):
        for key, value in my_dict.items():
             if val == value:
                 return key

    #Scrapes webpage for which teams played
    #inputs are a team dictionary the team we are looking at and the match num
    def scrape_match_teams_playing(self, teams, team_id, match_num):
        flag = 0
        team = teams.get(str(team_id))
        URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid=" + str(match_num)
        #print(URL)
        current_team = (teams[str(team_id)])
        page = self.requests.get(URL)
        soup = self.BeautifulSoup(page.content, 'html.parser')
        #returns the teams playing for each match by getting the text element
        #of the HTML returned by the soup object
        data = [element.text for element in soup.find_all('td', class_='bnorm', width='190')]
        #print(data)
        #match_data = results.find_all('td', class_="statdata")
        if current_team in data:
            flag = 1
        return flag


    #Scrapes webpage for the match stats and returns an array of the data
    #should also add the adv stats to the data
    #inputs the teams dict, current team
    #update id: if its 0 it means start from the beggining
    #if its a match_ID it only updates stats from a certain game in the text file
    def scrape_match_stats(self, teams, team_id, update_id, update_end):
        team = teams.get(str(team_id))
        f = open("Data/"+team+"_data.txt", 'r')
        M_IDs = f.readlines()
        M_IDs = [x.rstrip() for x in M_IDs]
        count = 1
        found = 0
        M_IDs_toappend = []
        for mn in M_IDs:
            if((update_id <= int(mn) and int(mn) <= update_end) or update_id == 0):
                print(mn + "made it into found")
                M_IDs_toappend.append(mn)

        for mn in M_IDs_toappend:
            print("into url area")
            #start the for loop here and keep track of line num
            URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid="+str(mn)
            advURL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid="+str(mn)+"&advv=Y"
            print(URL)
            page = self.requests.get(URL)
            advpage = self.requests.get(advURL)
            soup = self.BeautifulSoup(page.content, 'html.parser')
            adv_soup = self.BeautifulSoup(advpage.content, 'html.parser')
            stat_array = []
            stat_array.append(int(mn))
            yr = self.determine_year(mn)
            stat_array.append(int(yr))
            data = [element.text for element in soup.find_all('td', class_='bnorm', width='190')]
            round_data = [element.text for element in soup.find_all('td', class_="lnorm", height='22')]
            points_table = [element.text for element in soup.find_all('td', align='center')]
            home_score = points_table[5].strip()
            away_score = points_table[10].strip()
            #uses round_data to chop out which round the match is being played as
            round = self.determine_round(round_data[0])
            stat_array.append(round)
            #determines if the given team for the match is home or away
            home = 0
            #the home team appears first  in the data array. 0 represents home team
            if(team == data[0]):
                stat_array.append(0) #adds home value as 0 as current team is home team
                points_for = int(home_score) #assigns points for as the home score as current team is home
                points_against = int(away_score)
                margin = points_for-points_against
                win_value = self.determine_win(margin)
                if(win_value == 1):
                    stat_array.append(0) #Home team won
                elif(win_value == 0):
                    stat_array.append(1) #away team won
                else:
                    stat_array.append(0.5) #draw
                stat_array.extend([points_for, points_against, margin, win_value]) # appends all the above
                oppo_ID = self.get_key(data[1], teams)
                stat_array.append(int(oppo_ID))
            #otherwise they must be the away team. 1 represents away team
            else:
                home = 1
                stat_array.append(1) #adds home value as 1 as current team is away team
                points_for = int(away_score) #determines point for as the away score as current team is away
                points_against = int(home_score)
                margin = points_for-points_against
                win_value = self.determine_win(margin)
                if(win_value == 1):
                    stat_array.append(1) #away team won
                elif(win_value == 0):
                    stat_array.append(0) #home team won
                else:
                    stat_array.append(0.5) #draw
                stat_array.extend([points_for, points_against, margin, win_value]) #appends above
                oppo_ID = self.get_key(data[0], teams)
                stat_array.append(int(oppo_ID))
            stats_pulled = [element.text for element in soup.find_all('td', class_="statdata")]
            adv_stats_pulled = [element.text for element in adv_soup.find_all('td', class_="statdata")]
            stat_array = self.wrangle_stats(stats_pulled, adv_stats_pulled, stat_array, home)
            self.write_to_excel(team, stat_array, count)
            count = count + 1

    #takes match number and determines whether it what year that was
    def determine_year(self, x):
        x = int(x)
        year = 9999
        if(x >= 5147 and x <= 5342):
            year = 2011
        elif(x >= 5343 and x <= 5549):
            year = 2012
        elif(x >= 5550 and x <= 5756):
            year = 2013
        elif(x >= 5757 and x <= 5963):
            year = 2014
        elif(x >= 5964 and x <= 6171):
            year = 2015
        elif((x >= 6172 and x <= 6369) or (x >= 9298 and x <= 9306 )):
            year = 2016
        elif(x >= 9307 and x <= 9513):
            year = 2017
        elif(x >= 9514 and x <= 9720):
            year = 2018
        elif(x >= 9721 and x <= 9927):
            year = 2019
        elif(x >= 9928 and x <= 10326):
            year = 2020
        elif(x>=10327 and x <= 10543):
            year = 2021
        elif(x>10543):
            year = 2022
        return year

    #takes the big soup of all HTML text code labelled as 'statdata'
    #reverses the soup as the important stuff is at the back
    #kicks is the 'last stat' so it counts to kicks + 1 for the final stat point
    #splices all other data off and leaves us with the good stuff
    #if the team we are getting stats for is the home team it starts from array = 0
    #otherwise it starts from 2 and either way they do every third step
    def wrangle_stats(self, stats_pulled, adv_stats_pulled, stat_array, home):
        stats_pulled.reverse() #chops up the bits of basic stats we want
        adv_stats_pulled.reverse()
        i = 1
        j = 0
        for x in stats_pulled:
            if(x == "150 or more"): #ensures that the backof the array starts at the 150 games or more area
                j = i
            if(x == "Kicks"):
                i = i + 1
                break
            i = i+1
        if (j>0):
            j = j-2
            stats_pulled = stats_pulled[j:i]
        else:
            stats_pulled = stats_pulled[j:i]
        stats_pulled.reverse()
        o = 1
        p = 0
        for y in adv_stats_pulled:
            if(y == "% Goals Assisted"):
                p = o
            if (y == "Contested Possessions"):
                o = o + 1
                break
            o = o + 1
        p = p-2
        adv_stats_pulled = adv_stats_pulled[p:o]
        adv_stats_pulled.reverse()
        if(home == 0): #does the home stats first as home team is the'for' team
            for x in stats_pulled[::3]:
                x = self.strip_units(x)
                stat_array.append(float(x))
            for y in adv_stats_pulled[::3]: #adds home teams advanced stats after basic
                y = self.strip_units(y)
                stat_array.append(float(y))
            for x in stats_pulled[2::3]: #adds away stats as away is team 'against'
                x = self.strip_units(x)
                stat_array.append(float(x))
            for y in adv_stats_pulled[2::3]:
                y = self.strip_units(y)
                stat_array.append(float(y))
        #does the same as above except the current team is the away teams
        else: #does the away stats first as the away team is the against team
            for x in stats_pulled[2::3]:
                x = self.strip_units(x)
                stat_array.append(float(x))
            for y in adv_stats_pulled[2::3]:
                y = self.strip_units(y)
                stat_array.append(float(y))
            for x in stats_pulled[::3]: #adds home stats as home team is 'against'
                x = self.strip_units(x)
                stat_array.append(float(x))
            for y in adv_stats_pulled[::3]: #adds home teams advanced stats after basic
                y = self.strip_units(y)
                stat_array.append(float(y))
        return stat_array

    #takes a margin value and determines if the match was drawn, won or lost for the margin
    def determine_win(self, x):
        value = 999
        if(x>0):
            value = 1
        elif(x<0):
            value = 0
        elif(x == 0):
            value = 0.5
        return value

    #Gets rid of the non-numeric units in the stats
    def strip_units(self, x):
        #gets rid of % mark
        if("%" in x):
            x = x[:-1]
        #gets ride of kg or cm
        elif(("cm" in x) or ("kg" in x)):
            x = x[:-2]
        #gets rid of mth
        elif("mth" in x):
            x = x[:-3]
            #if mth is 10 or 11
            if(len(x) == 7):
                y = x[-1]+x[-2]
                #makes the mth = to decimal
                z = float(y)*8.33
                #don't think this would actually happen
                if(z < 9):
                    z = str(z)
                    z = z[:1]
                    x = x[:2] + '.' + z
                else:
                    z = str(z)
                    z = z[:2]
                    x = x[:2] + '.' + z
            #if mth is 0-9
            else:
                y = x[-1]
                z = float(y)*8.33
                #will be less than 9 if mth is 0 or 1
                #will result in turning 24yr 11mth to 24.91
                if(z < 9):
                    z = str(z)
                    z = z[:1]
                    x = x[:2] + '.' + z
                else:
                    z = str(z)
                    z = z[:2]
                    x = x[:2] + '.' + z
        return x

    #Example input 'Round 23, Marvel Stadium... etc' will look at either first character
    # or comma position to determine what round it is
    # finals rounds are given 25-28 values accordingly
    def determine_round(self, round_string):
        round = 0
        if (round_string[0] == 'Q' or round_string[0] == 'E'):
            round = 25
        elif (round_string[0] == "S"):
            round = 26
        elif (round_string[0] == "P"):
            round = 27
        elif (round_string[0] == "G"):
            round = 28
        else:
            if(round_string[7] == ","):
                round = int(round_string[6])
            else:
                round = int(round_string[6] + round_string[7])
        return round

    #opens an excel file based on the team name
    #if the excel file doesn't exist it creates the excel file and adds the labeled column
    #along with the first set of statistics
    #otherwise it opens the existing file and adds the relevant stats into the next open column
    def write_to_excel(self, team, stat_array, match_count):
        if(not(self.path.exists("Data/"+team+'_stats.xlsx'))):
            wb = self.Workbook()
            ws = wb.active
            labels = ['Match_ID', 'Year', 'Round', 'H/A?', 'H/A Win?', 'Points For', 'Points Against', 'Margin',
            'Team Won? (1=W, 0=L)', 'Team_against_ID', 'Kicks', 'Handballs', 'Disposals',
            'Kick to HB Ratio', 'Marks', 'Tackles', 'Hitouts', 'Frees For', 'Frees Against',
            'Goals Kicked', 'Goal Assists1', 'Behinds Kicked', 'Rushed Behinds', 'Scoring Shots',
            'Conversion %', 'Disposals Per Goal', 'Disposals Per Scoring Shot', 'Clearances1',
            'Clangers1', 'Rebound 50s1', 'Inside 50s', 'In50s Per Scoring Shot', 'In50s Per Goal',
            '% In50s Score', '% In50s Goal', 'Height', 'Weight', 'Age', 'Av Games', '<50 Games',
            '50-99 Games', '100-149 Games', '>150 Games', 'Contested Poss', 'Uncontested Poss', 'Effective Disposals',
            'Disposal Efficiency', 'Clangers2', 'Contested Marks', 'Marks Inside 50', 'Clearances2', 'Rebound 50s2',
            '1%ers', 'Bounces', 'Goals Assists2', 'Goal Assist %', 'Kicks (O)', 'Handballs (O)', 'Disposals (O)',
            'Kick to HB Ratio (O)', 'Marks (O)', 'Tackles(O)', 'Hitouts(O)', 'Frees For(O)', 'Frees Against(O)',
            'Goals Kicked(O)', 'Goal Assists1(O)', 'Behinds Kicked(O)', 'Rushed Behinds(O)', 'Scoring Shots(O)',
            'Conversion %(O)', 'Disposals Per Goal(O)', 'Disposals Per Scoring Shot(O)', 'Clearances1(O)',
            'Clangers1(O)', 'Rebound 50s1(O)', 'Inside 50s(O)', 'In50s Per Scoring Shot(O)', 'In50s Per Goal(O)',
            '% In50s Score(O)', '% In50s Goal(O)', 'Height(O)', 'Weight(O)', 'Age(O)', 'Av Games(O)', '<50 Games(O)',
            '50-99 Games(O)', '100-149 Games(O)', '>150 Games(O)', 'Contested Poss(O)', 'Uncontested Poss(O)', 'Effective Disposals(O)',
            'Disposal Efficiency(O)', 'Clangers2(O)', 'Contested Marks(O)', 'Marks Inside 50(O)', 'Clearances2(O)', 'Rebound 50s2(O)',
            '1%ers(O)', 'Bounces(O)', 'Goals Assists2(O)', 'Goal Assist %(O)']
            i = 0
            j = 0
            #iterates through each column, in the given range, here it is the 1st column
            #goes to maximum rows of the length of labels for stats
            for col in ws.iter_cols(max_col=1, max_row=len(labels)):
                for cell in col:
                    cell.value = labels[i]
                    i = i+1
            #iterates to the column that should be free next
            for col in ws.iter_cols(min_col=match_count+1, max_col=match_count+1, max_row=len(stat_array)):
                for cell in col:
                    cell.value = stat_array[j]
                    j = j + 1
            wb.save(filename = team+'_stats.xlsx')
        else:
            from openpyxl import Workbook, load_workbook
            wb = load_workbook("Data/"+team+'_stats.xlsx')
            ws = wb.active
            j = 0
            match_count = ws.max_column
            for col in ws.iter_cols(min_col=match_count+1, max_col=match_count+1, max_row=len(stat_array)):
                for cell in col:
                    cell.value = stat_array[j]
                    j = j + 1
            wb.save(filename = ("Data/"+team+'_stats.xlsx'))

    #creates a dictionary with each teams identifier on afl_tables
    def createTeamDict(self):
        teamDict = {
        "1" : "Adelaide",
        "2" : "Brisbane",
        "3" : "Carlton",
        "4" : "Collingwood",
        "5" : "Essendon",
        "6" : "Fremantle",
        "7" : "Geelong",
        "8" : "Gold Coast",
        "9" : "GWS",
        "10": "Hawthorn",
        "11": "Melbourne",
        "12": "North Melbourne",
        "13": "Port Adelaide",
        "14": "Richmond",
        "15": "St Kilda",
        "16": "Sydney",
        "17": "West Coast",
        "18": "Western Bulldogs"
        }
        return teamDict

    #creates a file for each team that has the ID's of each of their matches  since 2011
    def createTeamMatchFile(team_int, team_dict):
        #gets current team we are looking at
        current_team = (team_dict[str(team_int)])
        print(current_team)
        textfile = open("Data/"+current_team+"_data.txt", 'w+')
        match = 1
        #round 1 2011, start of the expansion teams
        p = 5147
        while(p<6370):
            #checks if current game has current team
            flag_num = scrape_match_teams_playing(team_dict, team_int, p)
            #if it does it records the match ID, makes it easier for the future
            if(flag_num == 1):
                #matches played because it doesn't take into account the bye round
                print("Match: "+str(match))
                match = match + 1
                textfile.write(str(p)+'\n')
            p = p+1
        #match starts at EF 2016 WC vs WB
        j = 9298
        #End of Round 1 2020
        while(j<9936):
            #checks if current game has current team
            flag_num = scrape_match_teams_playing(team_dict, team_int, j)
            #if it does it records the match ID, makes it easier for the future
            if(flag_num == 1):
                #matches played because it doesn't take into account the bye round
                print("match: "+str(match))
                match = match + 1
                textfile.write(str(j)+'\n')
            j = j+1
        textfile.close()

    def update_match_files(self, match_id, team_id, teams):
        flag_num = self.scrape_match_teams_playing(teams, team_id, match_id)
        if(flag_num == 1):
            current_team = (teams[str(team_id)])
            print("Team in match ID: " + str(match_id) + " is " + current_team)
            textfile = open("Data/"+current_team+"_data.txt", 'a')
            textfile.write(str(match_id)+"\n")

    #Updates locally stored stats update_match_files
    #Will go through a range of match ID's and put them in appropriate text files (x)
    #Then using x as the update value, will update the excel sheets from this value onwards
        #eg. if I wanted to update every match_ID from match 9000, i'd set x as 9000
    def update(self, M_ID_to_start_from, M_ID_to_end, teams):
        y = M_ID_to_end
        i = 1
        #updates match files from x to y
        while(i<19):
            x = M_ID_to_start_from
            while (x<=y):
                self.update_match_files(x,i,teams)
                x = x + 1
            i = i+1
        i = 1
        x = M_ID_to_start_from
        #adds the added matches into the excel spreadsheet from a minimum of x
        while(i<19):
            self.scrape_match_stats(teams, i, x, y)
            i = i+1

    def main():
    #    g = gatherer()
    #    teams = g.createTeamDict()
    #    g.update(int(sys.argv[1]), int(sys.argv[2]),teams)
        i = 18
        #should go through each of the 18 teams and create an excel file with over 100 stats for each game they played in
        while(i<19):
            scrape_match_stats(teams,i)
            clean_match_stats(teams,i)
            i = i+1

    if __name__ == '__main__':
        main()
